#!/usr/bin/env python3
"""
eSubpoena Tolls Tool - Professional Telecommunication Data Analysis
A Python application for parsing XML telecommunication data and exporting to Excel with analytics.
"""

import sys
import os
import re
import json
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QTextEdit, QProgressBar, QFileDialog, QMessageBox, QFrame,
    QGridLayout, QGroupBox, QSplitter, QScrollArea, QHeaderView
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QMimeData
from PyQt6.QtGui import QFont, QPalette, QColor, QDragEnterEvent, QDropEvent, QPixmap, QIcon

import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class CallRecord:
    """Represents a single call record with normalized data."""
    
    def __init__(self, direction: str, remote_number: str, start_time: str, 
                 end_time: str, length_of_call: int):
        self.direction = direction
        self.remote_number = remote_number
        self.normalized_number = self._normalize_phone_number(remote_number)
        self.start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        self.end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
        self.length_of_call = length_of_call
        self.duration_minutes = length_of_call / 60.0
        self.date = self.start_time.strftime('%Y-%m-%d')
        self.time = self.start_time.strftime('%H:%M:%S')
        self.day_of_week = self.start_time.strftime('%A')
    
    def _normalize_phone_number(self, number: str) -> str:
        """Normalize phone number to 10-digit format."""
        # Remove all non-digit characters
        digits_only = re.sub(r'[^\d]', '', str(number))
        
        # If it's a 10-digit number, return as is
        if len(digits_only) == 10:
            return digits_only
        
        # If it's 11 digits and starts with 1, remove the 1
        if len(digits_only) == 11 and digits_only.startswith('1'):
            return digits_only[1:]
        
        # If it's longer than 10 digits, take the last 10
        if len(digits_only) > 10:
            return digits_only[-10:]
        
        # Otherwise, pad with zeros to make it 10 digits
        return digits_only.zfill(10)


class Analytics:
    """Handles analytics calculations for call records."""
    
    def __init__(self, records: List[CallRecord]):
        self.records = records
        self._calculate_analytics()
    
    def _calculate_analytics(self):
        """Calculate all analytics from the call records."""
        if not self.records:
            self.total_calls = 0
            self.incoming_calls = 0
            self.outgoing_calls = 0
            self.total_duration_minutes = 0.0
            self.average_call_duration = 0.0
            self.unique_numbers = 0
            self.most_frequent_numbers = []
            self.calls_by_day = {}
            self.calls_by_hour = {}
            self.longest_call = None
            self.shortest_call = None
            return
        
        self.total_calls = len(self.records)
        self.incoming_calls = sum(1 for r in self.records if r.direction.lower() == 'incoming')
        self.outgoing_calls = self.total_calls - self.incoming_calls
        
        self.total_duration_minutes = sum(r.duration_minutes for r in self.records)
        self.average_call_duration = self.total_duration_minutes / self.total_calls if self.total_calls > 0 else 0
        
        # Unique numbers
        unique_numbers = set(r.normalized_number for r in self.records)
        self.unique_numbers = len(unique_numbers)
        
        # Most frequent numbers
        number_counts = {}
        for record in self.records:
            number_counts[record.normalized_number] = number_counts.get(record.normalized_number, 0) + 1
        
        self.most_frequent_numbers = sorted(number_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # Calls by day
        self.calls_by_day = {}
        for record in self.records:
            self.calls_by_day[record.date] = self.calls_by_day.get(record.date, 0) + 1
        
        # Calls by hour
        self.calls_by_hour = {}
        for record in self.records:
            hour = record.start_time.hour
            self.calls_by_hour[hour] = self.calls_by_hour.get(hour, 0) + 1
        
        # Longest and shortest calls
        if self.records:
            self.longest_call = max(self.records, key=lambda r: r.length_of_call)
            self.shortest_call = min((r for r in self.records if r.length_of_call > 0), 
                                   key=lambda r: r.length_of_call, default=None)


class XMLParser:
    """Handles XML parsing for telecommunication data."""
    
    @staticmethod
    def parse_file(file_path: str) -> List[CallRecord]:
        """Parse XML file and extract call records."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Find all results elements
            results = root.findall('.//results')
            records = []
            
            for result in results:
                try:
                    direction = result.find('messageDirection').text
                    remote_number = result.find('remoteNumber').text
                    start_time = result.find('startTime').text
                    end_time = result.find('endTime').text
                    length_of_call = int(result.find('lengthOfCall').text)
                    
                    record = CallRecord(direction, remote_number, start_time, end_time, length_of_call)
                    records.append(record)
                    
                except (AttributeError, ValueError) as e:
                    print(f"Error processing record: {e}")
                    continue
            
            return records
            
        except Exception as e:
            raise Exception(f"Failed to parse XML file: {e}")


class ExcelExporter:
    """Handles Excel export with professional formatting."""
    
    @staticmethod
    def export_data(records: List[CallRecord], analytics: Analytics, output_path: str):
        """Export data to Excel with professional formatting."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        ws_records = wb.create_sheet("Call Records")
        ws_analytics = wb.create_sheet("Analytics")
        ws_summary = wb.create_sheet("Summary Report")
        
        # Export call records
        ExcelExporter._export_call_records(ws_records, records)
        
        # Export analytics
        ExcelExporter._export_analytics(ws_analytics, analytics)
        
        # Export summary report
        ExcelExporter._export_summary_report(ws_summary, analytics)
        
        # Save workbook
        wb.save(output_path)
    
    @staticmethod
    def _export_call_records(ws, records: List[CallRecord]):
        """Export call records to worksheet."""
        # Headers
        headers = [
            "Direction", "Remote Number", "Normalized Number", "Start Time", "End Time",
            "Duration (sec)", "Duration (min)", "Date", "Time", "Day of Week"
        ]
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.direction).border = border
            ws.cell(row=row, column=2, value=record.remote_number).border = border
            ws.cell(row=row, column=3, value=record.normalized_number).border = border
            ws.cell(row=row, column=4, value=record.start_time).border = border
            ws.cell(row=row, column=5, value=record.end_time).border = border
            ws.cell(row=row, column=6, value=record.length_of_call).border = border
            ws.cell(row=row, column=7, value=record.duration_minutes).border = border
            ws.cell(row=row, column=8, value=record.date).border = border
            ws.cell(row=row, column=9, value=record.time).border = border
            ws.cell(row=row, column=10, value=record.day_of_week).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _export_analytics(ws, analytics: Analytics):
        """Export analytics to worksheet."""
        # Summary statistics
        summary_data = [
            ("Total Calls", analytics.total_calls),
            ("Incoming Calls", analytics.incoming_calls),
            ("Outgoing Calls", analytics.outgoing_calls),
            ("Unique Phone Numbers", analytics.unique_numbers),
            ("Total Duration (minutes)", f"{analytics.total_duration_minutes:.2f}"),
            ("Average Call Duration (minutes)", f"{analytics.average_call_duration:.2f}"),
        ]
        
        # Style
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write summary
        ws.cell(row=1, column=1, value="Metric").font = header_font
        ws.cell(row=1, column=2, value="Value").font = header_font
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws.cell(row=row, column=1, value=metric).border = border
            ws.cell(row=row, column=2, value=value).border = border
        
        # Most frequent numbers
        start_row = len(summary_data) + 4
        ws.cell(row=start_row, column=1, value="Most Frequent Numbers").font = header_font
        ws.cell(row=start_row, column=2, value="Call Count").font = header_font
        
        for i, (number, count) in enumerate(analytics.most_frequent_numbers, 1):
            ws.cell(row=start_row + i, column=1, value=number).border = border
            ws.cell(row=start_row + i, column=2, value=count).border = border
        
        # Calls by day
        day_start_row = start_row + len(analytics.most_frequent_numbers) + 3
        ws.cell(row=day_start_row, column=1, value="Calls by Day").font = header_font
        ws.cell(row=day_start_row, column=2, value="Call Count").font = header_font
        
        sorted_days = sorted(analytics.calls_by_day.items())
        for i, (day, count) in enumerate(sorted_days, 1):
            ws.cell(row=day_start_row + i, column=1, value=day).border = border
            ws.cell(row=day_start_row + i, column=2, value=count).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _export_summary_report(ws, analytics: Analytics):
        """Export summary report to worksheet."""
        report_lines = [
            "=== TELECOMMUNICATION DATA ANALYSIS ===",
            "",
            f"Total Calls: {analytics.total_calls}",
            f"Incoming Calls: {analytics.incoming_calls}",
            f"Outgoing Calls: {analytics.outgoing_calls}",
            f"Unique Phone Numbers: {analytics.unique_numbers}",
            f"Total Duration: {analytics.total_duration_minutes:.2f} minutes",
            f"Average Call Duration: {analytics.average_call_duration:.2f} minutes",
            "",
            "=== MOST FREQUENT NUMBERS ===",
        ]
        
        for i, (number, count) in enumerate(analytics.most_frequent_numbers, 1):
            report_lines.append(f"{i}. {number} ({count} calls)")
        
        report_lines.extend([
            "",
            "=== CALLS BY DAY ===",
        ])
        
        sorted_days = sorted(analytics.calls_by_day.items())
        for day, count in sorted_days:
            report_lines.append(f"{day}: {count} calls")
        
        # Write report
        for row, line in enumerate(report_lines, 1):
            cell = ws.cell(row=row, column=1, value=line)
            if line.startswith("==="):
                cell.font = Font(bold=True)
        
        # Set column width
        ws.column_dimensions['A'].width = 80


class ProcessingThread(QThread):
    """Background thread for processing XML files."""
    
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, file_path: str):
        super().__init__()
        self.file_path = file_path
    
    def run(self):
        try:
            self.progress.emit("Parsing XML file...")
            records = XMLParser.parse_file(self.file_path)
            self.progress.emit(f"Found {len(records)} call records")
            self.finished.emit(records)
        except Exception as e:
            self.error.emit(str(e))


class DragDropWidget(QWidget):
    """Custom widget for drag and drop functionality."""
    
    file_dropped = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setMinimumHeight(200)
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                border: 2px dashed #cccccc;
                border-radius: 10px;
            }
            QWidget:hover {
                border-color: #0078d4;
                background-color: #e6f3ff;
            }
        """)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Icon and text
        self.label = QLabel("📁 Drag and drop XML files here\nor click to browse")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setStyleSheet("font-size: 16px; color: #666666;")
        layout.addWidget(self.label)
        
        # Browse button
        self.browse_btn = QPushButton("Browse Files")
        self.browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
        """)
        self.browse_btn.clicked.connect(self._browse_files)
        layout.addWidget(self.browse_btn, alignment=Qt.AlignmentFlag.AlignCenter)
    
    def _browse_files(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select XML File", "", "XML Files (*.xml)"
        )
        if file_path:
            self.file_dropped.emit(file_path)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QWidget {
                    background-color: #e6f3ff;
                    border: 2px dashed #0078d4;
                    border-radius: 10px;
                }
            """)
    
    def dragLeaveEvent(self, event):
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                border: 2px dashed #cccccc;
                border-radius: 10px;
            }
            QWidget:hover {
                border-color: #0078d4;
                background-color: #e6f3ff;
            }
        """)
    
    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                border: 2px dashed #cccccc;
                border-radius: 10px;
            }
            QWidget:hover {
                border-color: #0078d4;
                background-color: #e6f3ff;
            }
        """)
        
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith('.xml'):
                self.file_dropped.emit(file_path)
                break


class EsubpoenaTollsTool(QMainWindow):
    """Main application window."""
    
    def __init__(self):
        super().__init__()
        self.records = []
        self.analytics = None
        self.processing_thread = None
        
        self.setWindowTitle("📞 eSubpoena Tolls Tool")
        self.setGeometry(100, 100, 1400, 900)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QTabWidget::pane {
                border: 1px solid #cccccc;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e0e0e0;
                padding: 10px 20px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #0078d4;
                color: white;
            }
            QTableWidget {
                gridline-color: #cccccc;
                selection-background-color: #e6f3ff;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 8px;
                border: 1px solid #cccccc;
                font-weight: bold;
            }
        """)
        
        self._setup_ui()
    
    def _setup_ui(self):
        """Setup the user interface."""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        # Header
        header = QHBoxLayout()
        title = QLabel("📞 eSubpoena Tolls Tool")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #0078d4;")
        header.addWidget(title)
        
        header.addStretch()
        
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #107c10;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0e6e0e;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.export_btn.clicked.connect(self._export_to_excel)
        self.export_btn.setEnabled(False)
        header.addWidget(self.export_btn)
        
        layout.addLayout(header)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Tab widget
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # Create tabs
        self._create_overview_tab()
        self._create_records_tab()
        self._create_analytics_tab()
        self._create_summary_tab()
    
    def _create_overview_tab(self):
        """Create the overview tab."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Welcome message
        welcome = QLabel("Welcome to eSubpoena Tolls Tool")
        welcome.setStyleSheet("font-size: 20px; font-weight: bold; margin: 20px;")
        welcome.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(welcome)
        
        description = QLabel("This tool processes telecommunication XML data and provides comprehensive analytics.")
        description.setStyleSheet("font-size: 14px; color: #666666; margin: 10px;")
        description.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(description)
        
        layout.addSpacing(20)
        
        # Drag and drop area
        self.drag_drop_widget = DragDropWidget()
        self.drag_drop_widget.file_dropped.connect(self._process_file)
        layout.addWidget(self.drag_drop_widget)
        
        layout.addSpacing(20)
        
        # Status label
        self.status_label = QLabel("Ready to process files")
        self.status_label.setStyleSheet("font-size: 14px; color: #666666;")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Statistics frame
        self.stats_frame = QFrame()
        self.stats_frame.setVisible(False)
        self.stats_frame.setStyleSheet("""
            QFrame {
                border: 1px solid #cccccc;
                border-radius: 10px;
                background-color: white;
                padding: 20px;
            }
        """)
        stats_layout = QVBoxLayout(self.stats_frame)
        
        stats_title = QLabel("Quick Statistics")
        stats_title.setStyleSheet("font-size: 18px; font-weight: bold; margin-bottom: 10px;")
        stats_layout.addWidget(stats_title)
        
        self.stats_grid = QGridLayout()
        stats_layout.addLayout(self.stats_grid)
        
        layout.addWidget(self.stats_frame)
        
        layout.addStretch()
        
        self.tab_widget.addTab(widget, "Overview")
    
    def _create_records_tab(self):
        """Create the call records tab."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Header
        header = QHBoxLayout()
        self.records_label = QLabel("No call records loaded")
        header.addWidget(self.records_label)
        
        header.addStretch()
        
        self.export_records_btn = QPushButton("Export to Excel")
        self.export_records_btn.setStyleSheet("""
            QPushButton {
                background-color: #107c10;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #0e6e0e;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.export_records_btn.clicked.connect(self._export_to_excel)
        self.export_records_btn.setEnabled(False)
        header.addWidget(self.export_records_btn)
        
        layout.addLayout(header)
        
        # Table
        self.records_table = QTableWidget()
        self.records_table.setColumnCount(10)
        self.records_table.setHorizontalHeaderLabels([
            "Direction", "Remote Number", "Normalized", "Date", "Time", 
            "Duration (min)", "Day of Week", "Start Time", "End Time", "Duration (sec)"
        ])
        
        # Set table properties
        self.records_table.setAlternatingRowColors(True)
        self.records_table.horizontalHeader().setStretchLastSection(True)
        self.records_table.setSortingEnabled(True)
        
        layout.addWidget(self.records_table)
        
        self.tab_widget.addTab(widget, "Call Records")
    
    def _create_analytics_tab(self):
        """Create the analytics tab."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Analytics title
        title = QLabel("Analytics Dashboard")
        title.setStyleSheet("font-size: 20px; font-weight: bold; margin: 10px;")
        layout.addWidget(title)
        
        # Statistics cards
        self.stats_cards = QHBoxLayout()
        layout.addLayout(self.stats_cards)
        
        # Analytics content
        self.analytics_content = QWidget()
        self.analytics_content.setVisible(False)
        analytics_layout = QVBoxLayout(self.analytics_content)
        
        # Most frequent numbers
        self.frequent_numbers_table = QTableWidget()
        self.frequent_numbers_table.setColumnCount(3)
        self.frequent_numbers_table.setHorizontalHeaderLabels(["Rank", "Phone Number", "Call Count"])
        analytics_layout.addWidget(QLabel("Most Frequent Numbers"))
        analytics_layout.addWidget(self.frequent_numbers_table)
        
        # Calls by day
        self.calls_by_day_table = QTableWidget()
        self.calls_by_day_table.setColumnCount(2)
        self.calls_by_day_table.setHorizontalHeaderLabels(["Date", "Call Count"])
        analytics_layout.addWidget(QLabel("Calls by Day"))
        analytics_layout.addWidget(self.calls_by_day_table)
        
        layout.addWidget(self.analytics_content)
        
        # No data message
        self.no_data_label = QLabel("No analytics available. Please process an XML file first.")
        self.no_data_label.setStyleSheet("font-size: 16px; color: #666666;")
        self.no_data_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.no_data_label)
        
        layout.addStretch()
        
        self.tab_widget.addTab(widget, "Analytics")
    
    def _create_summary_tab(self):
        """Create the summary tab."""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Header
        header = QHBoxLayout()
        header.addWidget(QLabel("Summary Report"))
        
        header.addStretch()
        
        self.copy_btn = QPushButton("Copy to Clipboard")
        self.copy_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #106ebe;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.copy_btn.clicked.connect(self._copy_to_clipboard)
        self.copy_btn.setEnabled(False)
        header.addWidget(self.copy_btn)
        
        layout.addLayout(header)
        
        # Summary text
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        self.summary_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #cccccc;
                border-radius: 5px;
                padding: 10px;
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: 12px;
            }
        """)
        layout.addWidget(self.summary_text)
        
        self.tab_widget.addTab(widget, "Summary")
    
    def _process_file(self, file_path: str):
        """Process the dropped XML file."""
        self.status_label.setText("Processing file...")
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        
        # Start processing thread
        self.processing_thread = ProcessingThread(file_path)
        self.processing_thread.progress.connect(self.status_label.setText)
        self.processing_thread.finished.connect(self._on_processing_finished)
        self.processing_thread.error.connect(self._on_processing_error)
        self.processing_thread.start()
    
    def _on_processing_finished(self, records: List[CallRecord]):
        """Handle processing completion."""
        self.records = records
        self.analytics = Analytics(records)
        
        self.progress_bar.setVisible(False)
        self.status_label.setText(f"Successfully processed {len(records)} call records")
        
        # Update UI
        self._update_records_table()
        self._update_analytics()
        self._update_summary()
        self._update_statistics()
        
        # Enable export button
        self.export_btn.setEnabled(True)
        self.export_records_btn.setEnabled(True)
        self.copy_btn.setEnabled(True)
        
        # Switch to records tab
        self.tab_widget.setCurrentIndex(1)
    
    def _on_processing_error(self, error_message: str):
        """Handle processing error."""
        self.progress_bar.setVisible(False)
        self.status_label.setText(f"Error: {error_message}")
        QMessageBox.critical(self, "Error", f"Failed to process file:\n{error_message}")
    
    def _update_records_table(self):
        """Update the records table."""
        self.records_table.setRowCount(len(self.records))
        
        for row, record in enumerate(self.records):
            self.records_table.setItem(row, 0, QTableWidgetItem(record.direction))
            self.records_table.setItem(row, 1, QTableWidgetItem(record.remote_number))
            self.records_table.setItem(row, 2, QTableWidgetItem(record.normalized_number))
            self.records_table.setItem(row, 3, QTableWidgetItem(record.date))
            self.records_table.setItem(row, 4, QTableWidgetItem(record.time))
            self.records_table.setItem(row, 5, QTableWidgetItem(f"{record.duration_minutes:.2f}"))
            self.records_table.setItem(row, 6, QTableWidgetItem(record.day_of_week))
            self.records_table.setItem(row, 7, QTableWidgetItem(record.start_time.strftime('%Y-%m-%d %H:%M:%S')))
            self.records_table.setItem(row, 8, QTableWidgetItem(record.end_time.strftime('%Y-%m-%d %H:%M:%S')))
            self.records_table.setItem(row, 9, QTableWidgetItem(str(record.length_of_call)))
        
        self.records_label.setText(f"Showing {len(self.records)} call records")
        
        # Auto-resize columns
        self.records_table.resizeColumnsToContents()
    
    def _update_analytics(self):
        """Update the analytics tab."""
        if not self.analytics:
            return
        
        # Update statistics cards
        self._update_statistics_cards()
        
        # Update frequent numbers table
        self.frequent_numbers_table.setRowCount(len(self.analytics.most_frequent_numbers))
        for row, (number, count) in enumerate(self.analytics.most_frequent_numbers):
            self.frequent_numbers_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            self.frequent_numbers_table.setItem(row, 1, QTableWidgetItem(number))
            self.frequent_numbers_table.setItem(row, 2, QTableWidgetItem(str(count)))
        
        # Update calls by day table
        sorted_days = sorted(self.analytics.calls_by_day.items())
        self.calls_by_day_table.setRowCount(len(sorted_days))
        for row, (day, count) in enumerate(sorted_days):
            self.calls_by_day_table.setItem(row, 0, QTableWidgetItem(day))
            self.calls_by_day_table.setItem(row, 1, QTableWidgetItem(str(count)))
        
        # Auto-resize columns
        self.frequent_numbers_table.resizeColumnsToContents()
        self.calls_by_day_table.resizeColumnsToContents()
        
        # Show analytics content
        self.analytics_content.setVisible(True)
        self.no_data_label.setVisible(False)
    
    def _update_statistics_cards(self):
        """Update the statistics cards."""
        # Clear existing cards
        for i in reversed(range(self.stats_cards.count())):
            self.stats_cards.itemAt(i).widget().setParent(None)
        
        if not self.analytics:
            return
        
        # Create new cards
        stats = [
            ("Total Calls", str(self.analytics.total_calls), "📞"),
            ("Incoming", str(self.analytics.incoming_calls), "📥"),
            ("Outgoing", str(self.analytics.outgoing_calls), "📤"),
            ("Unique Numbers", str(self.analytics.unique_numbers), "👥"),
            ("Total Duration", f"{self.analytics.total_duration_minutes:.1f} min", "⏱️"),
            ("Avg Duration", f"{self.analytics.average_call_duration:.1f} min", "📊"),
        ]
        
        for title, value, icon in stats:
            card = self._create_stat_card(title, value, icon)
            self.stats_cards.addWidget(card)
    
    def _create_stat_card(self, title: str, value: str, icon: str) -> QFrame:
        """Create a statistics card."""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                border: 1px solid #cccccc;
                border-radius: 10px;
                background-color: white;
                padding: 15px;
                margin: 5px;
            }
        """)
        card.setFixedSize(200, 100)
        
        layout = QVBoxLayout(card)
        
        # Icon and title
        header = QHBoxLayout()
        icon_label = QLabel(icon)
        icon_label.setStyleSheet("font-size: 20px;")
        header.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 12px; color: #666666;")
        header.addWidget(title_label)
        header.addStretch()
        
        layout.addLayout(header)
        
        # Value
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #0078d4;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_label)
        
        return card
    
    def _update_statistics(self):
        """Update the overview statistics."""
        if not self.analytics:
            return
        
        # Clear existing stats
        for i in reversed(range(self.stats_grid.count())):
            self.stats_grid.itemAt(i).widget().setParent(None)
        
        # Add new stats
        stats = [
            ("Total Calls", str(self.analytics.total_calls)),
            ("Incoming", str(self.analytics.incoming_calls)),
            ("Outgoing", str(self.analytics.outgoing_calls)),
            ("Unique Numbers", str(self.analytics.unique_numbers)),
            ("Total Duration", f"{self.analytics.total_duration_minutes:.1f} min"),
            ("Avg Duration", f"{self.analytics.average_call_duration:.1f} min"),
        ]
        
        for i, (label, value) in enumerate(stats):
            row = i // 2
            col = (i % 2) * 2
            
            label_widget = QLabel(label)
            label_widget.setStyleSheet("font-weight: bold;")
            self.stats_grid.addWidget(label_widget, row, col)
            
            value_widget = QLabel(value)
            self.stats_grid.addWidget(value_widget, row, col + 1)
        
        self.stats_frame.setVisible(True)
    
    def _update_summary(self):
        """Update the summary report."""
        if not self.analytics:
            return
        
        report_lines = [
            "=== TELECOMMUNICATION DATA ANALYSIS ===\n",
            f"Total Calls: {self.analytics.total_calls}",
            f"Incoming Calls: {self.analytics.incoming_calls}",
            f"Outgoing Calls: {self.analytics.outgoing_calls}",
            f"Unique Phone Numbers: {self.analytics.unique_numbers}",
            f"Total Duration: {self.analytics.total_duration_minutes:.2f} minutes",
            f"Average Call Duration: {self.analytics.average_call_duration:.2f} minutes\n",
            "=== MOST FREQUENT NUMBERS ===",
        ]
        
        for i, (number, count) in enumerate(self.analytics.most_frequent_numbers, 1):
            report_lines.append(f"{i}. {number} ({count} calls)")
        
        report_lines.extend([
            "\n=== CALLS BY DAY ===",
        ])
        
        sorted_days = sorted(self.analytics.calls_by_day.items())
        for day, count in sorted_days:
            report_lines.append(f"{day}: {count} calls")
        
        self.summary_text.setPlainText("\n".join(report_lines))
    
    def _export_to_excel(self):
        """Export data to Excel."""
        if not self.records or not self.analytics:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "telecommunication_analysis.xlsx", 
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                ExcelExporter.export_data(self.records, self.analytics, file_path)
                QMessageBox.information(
                    self, "Success", 
                    f"Successfully exported to:\n{file_path}"
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", 
                    f"Export failed:\n{str(e)}"
                )
    
    def _copy_to_clipboard(self):
        """Copy summary report to clipboard."""
        if self.summary_text.toPlainText():
            clipboard = QApplication.clipboard()
            clipboard.setText(self.summary_text.toPlainText())
            QMessageBox.information(self, "Success", "Report copied to clipboard")


def main():
    """Main application entry point."""
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("eSubpoena Tolls Tool")
    app.setApplicationVersion("1.0.0")
    app.setOrganizationName("eSubpoena")
    
    # Create and show main window
    window = EsubpoenaTollsTool()
    window.show()
    
    # Run application
    sys.exit(app.exec())


if __name__ == "__main__":
    main() 