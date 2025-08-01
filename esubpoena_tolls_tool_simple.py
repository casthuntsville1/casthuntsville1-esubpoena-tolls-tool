#!/usr/bin/env python3
"""
eSubpoena Tolls Tool - Professional Telecommunication Data Analysis
A Python application for parsing XML telecommunication data and exporting to Excel with analytics.
Simplified version using tkinter (built-in with Python).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import re
import json
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from pathlib import Path

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class CallRecord:
    """Represents a single call record with normalized data."""
    
    def __init__(self, direction: str, remote_number: str, start_time: str, 
                 end_time: str, length_of_call: int, target_number: str = "", source_file: str = ""):
        self.direction = direction
        self.remote_number = remote_number
        self.normalized_number = self._normalize_phone_number(remote_number)
        self.target_number = self._normalize_phone_number(target_number) if target_number else ""
        self.source_file = source_file  # Track which XML file this record came from
        # Parse datetime and remove timezone info for Excel compatibility
        self.start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00')).replace(tzinfo=None)
        self.end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00')).replace(tzinfo=None)
        self.length_of_call = length_of_call
        self.duration_minutes = length_of_call / 60.0
        self.date = self.start_time.strftime('%Y-%m-%d')
        self.time = self.start_time.strftime('%H:%M:%S')
        self.date_time = self.start_time.strftime('%Y-%m-%d %H:%M:%S')  # Combined date and time
        self.day_of_week = self.start_time.strftime('%A')
    
    def _normalize_phone_number(self, number: str) -> str:
        """Normalize phone number to 10-digit format."""
        # Remove all non-digit characters
        digits_only = re.sub(r'[^\d]', '', str(number))
        
        # If it's a 10-digit number, return as is
        if len(digits_only) == 10:
            return digits_only
        
        # If it's 11 digits and starts with 1, remove the 1
        if len(digits_only) == 11 and digits_only.startswith('1'):
            return digits_only[1:]
        
        # If it's longer than 10 digits, take the last 10
        if len(digits_only) > 10:
            return digits_only[-10:]
        
        # Otherwise, pad with zeros to make it 10 digits
        return digits_only.zfill(10)


class Analytics:
    """Handles analytics calculations for call records."""
    
    def __init__(self, records: List[CallRecord]):
        self.records = records
        self._calculate_analytics()
    
    def _calculate_analytics(self):
        """Calculate all analytics from the call records."""
        if not self.records:
            self.total_calls = 0
            self.incoming_calls = 0
            self.outgoing_calls = 0
            self.total_duration_minutes = 0.0
            self.average_call_duration = 0.0
            self.unique_numbers = 0
            self.most_frequent_numbers = []
            self.calls_by_day = {}
            self.calls_by_hour = {}
            self.longest_call = None
            self.shortest_call = None
            self.target_numbers = set()
            self.common_contacts = []
            self.files_processed = set()
            return
        
        self.total_calls = len(self.records)
        self.incoming_calls = sum(1 for r in self.records if r.direction.lower() == 'incoming')
        self.outgoing_calls = self.total_calls - self.incoming_calls
        
        self.total_duration_minutes = sum(r.duration_minutes for r in self.records)
        self.average_call_duration = self.total_duration_minutes / self.total_calls if self.total_calls > 0 else 0
        
        # Track target numbers and source files
        self.target_numbers = set(r.target_number for r in self.records if r.target_number)
        self.files_processed = set(r.source_file for r in self.records if r.source_file)
        
        # Unique numbers
        unique_numbers = set(r.normalized_number for r in self.records)
        self.unique_numbers = len(unique_numbers)
        
        # Most frequent numbers
        number_counts = {}
        for record in self.records:
            number_counts[record.normalized_number] = number_counts.get(record.normalized_number, 0) + 1
        
        self.most_frequent_numbers = sorted(number_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # Find common contacts across different target numbers
        self._find_common_contacts()
        
        # Calls by day
        self.calls_by_day = {}
        for record in self.records:
            self.calls_by_day[record.date] = self.calls_by_day.get(record.date, 0) + 1
        
        # Calls by hour
        self.calls_by_hour = {}
        for record in self.records:
            hour = record.start_time.hour
            self.calls_by_hour[hour] = self.calls_by_hour.get(hour, 0) + 1
        
        # Longest and shortest calls
        if self.records:
            self.longest_call = max(self.records, key=lambda r: r.length_of_call)
            self.shortest_call = min((r for r in self.records if r.length_of_call > 0), 
                                   key=lambda r: r.length_of_call, default=None)
    
    def _find_common_contacts(self):
        """Find contacts that appear across multiple target numbers."""
        # Group records by target number
        target_groups = {}
        for record in self.records:
            if record.target_number:
                if record.target_number not in target_groups:
                    target_groups[record.target_number] = set()
                target_groups[record.target_number].add(record.normalized_number)
        
        # Find numbers that appear in multiple target groups
        if len(target_groups) > 1:
            all_numbers = set()
            for numbers in target_groups.values():
                all_numbers.update(numbers)
            
            self.common_contacts = []
            for number in all_numbers:
                target_numbers_with_contact = [target for target, numbers in target_groups.items() if number in numbers]
                if len(target_numbers_with_contact) > 1:
                    self.common_contacts.append({
                        'number': number,
                        'target_numbers': target_numbers_with_contact,
                        'count': len(target_numbers_with_contact)
                    })
            
            # Sort by number of target numbers they appear in
            self.common_contacts.sort(key=lambda x: x['count'], reverse=True)


class XMLParser:
    """Handles XML parsing for telecommunication data."""
    
    @staticmethod
    def parse_file(file_path: str) -> List[CallRecord]:
        """Parse XML file and extract call records."""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Find target number from the XML structure
            target_number = ""
            target_value_elem = root.find('.//targetValue')
            if target_value_elem is not None and target_value_elem.text:
                target_number = target_value_elem.text
            
            # Find all results elements
            results = root.findall('.//results')
            records = []
            
            for result in results:
                try:
                    direction = result.find('messageDirection').text
                    remote_number = result.find('remoteNumber').text
                    start_time = result.find('startTime').text
                    end_time = result.find('endTime').text
                    length_of_call = int(result.find('lengthOfCall').text)
                    
                    record = CallRecord(direction, remote_number, start_time, end_time, length_of_call, target_number, os.path.basename(file_path))
                    records.append(record)
                    
                except (AttributeError, ValueError) as e:
                    print(f"Error processing record: {e}")
                    continue
            
            return records
            
        except Exception as e:
            raise Exception(f"Failed to parse XML file: {e}")


class ExcelExporter:
    """Handles Excel export with professional formatting."""
    
    @staticmethod
    def export_data(records: List[CallRecord], analytics: Analytics, output_path: str):
        """Export data to Excel with professional formatting."""
        try:
            # Create a new workbook with proper data types
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create worksheets
            ws_records = wb.create_sheet("Call Records")
            ws_analytics = wb.create_sheet("Analytics")
            ws_summary = wb.create_sheet("Summary Report")
            ws_common = wb.create_sheet("Common Contacts")
            
            # Export call records
            ExcelExporter._export_call_records(ws_records, records)
            
            # Export analytics
            ExcelExporter._export_analytics(ws_analytics, analytics)
            
            # Export summary report
            ExcelExporter._export_summary_report(ws_summary, analytics)
            
            # Export common contacts
            ExcelExporter._export_common_contacts(ws_common, analytics)
            
            # Clean up any potential formula issues by ensuring all cells are properly formatted
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            # Ensure the cell value is properly formatted
                            if isinstance(cell.value, (int, float)):
                                # Keep numeric values as they are
                                pass
                            elif isinstance(cell.value, str):
                                # Ensure strings don't start with = (which could be interpreted as formulas)
                                if cell.value.startswith('='):
                                    cell.value = "'" + cell.value
            
            # Save workbook
            wb.save(output_path)
            
        except Exception as e:
            raise Exception(f"Excel export failed: {str(e)}")
    
    @staticmethod
    def _export_call_records(ws, records: List[CallRecord]):
        """Export call records to worksheet."""
        # Headers
        headers = [
            "Direction", "Target Number", "Remote Number", "Normalized Number", "Date & Time", "End Time",
            "Duration (sec)", "Duration (min)", "Day of Week", "Source File"
        ]
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, record in enumerate(records, 2):
            # Use proper data types to avoid formula issues
            ws.cell(row=row, column=1, value=record.direction).border = border
            ws.cell(row=row, column=2, value=record.target_number).border = border
            ws.cell(row=row, column=3, value=record.remote_number).border = border
            ws.cell(row=row, column=4, value=record.normalized_number).border = border
            ws.cell(row=row, column=5, value=record.date_time).border = border
            ws.cell(row=row, column=6, value=record.end_time.strftime('%Y-%m-%d %H:%M:%S')).border = border
            ws.cell(row=row, column=7, value=record.length_of_call).border = border
            ws.cell(row=row, column=8, value=record.duration_minutes).border = border
            ws.cell(row=row, column=9, value=record.day_of_week).border = border
            ws.cell(row=row, column=10, value=record.source_file).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _export_analytics(ws, analytics: Analytics):
        """Export analytics to worksheet."""
        # Summary statistics
        summary_data = [
            ("Total Calls", analytics.total_calls),
            ("Incoming Calls", analytics.incoming_calls),
            ("Outgoing Calls", analytics.outgoing_calls),
            ("Unique Phone Numbers", analytics.unique_numbers),
            ("Total Duration (minutes)", f"{analytics.total_duration_minutes:.2f}"),
            ("Average Call Duration (minutes)", f"{analytics.average_call_duration:.2f}"),
        ]
        
        # Style
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write summary
        ws.cell(row=1, column=1, value="Metric").font = header_font
        ws.cell(row=1, column=2, value="Value").font = header_font
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws.cell(row=row, column=1, value=metric).border = border
            ws.cell(row=row, column=2, value=value).border = border
        
        # Most frequent numbers
        start_row = len(summary_data) + 4
        ws.cell(row=start_row, column=1, value="Most Frequent Numbers").font = header_font
        ws.cell(row=start_row, column=2, value="Call Count").font = header_font
        
        for i, (number, count) in enumerate(analytics.most_frequent_numbers, 1):
            ws.cell(row=start_row + i, column=1, value=number).border = border
            ws.cell(row=start_row + i, column=2, value=count).border = border
        
        # Calls by day
        day_start_row = start_row + len(analytics.most_frequent_numbers) + 3
        ws.cell(row=day_start_row, column=1, value="Calls by Day").font = header_font
        ws.cell(row=day_start_row, column=2, value="Call Count").font = header_font
        
        sorted_days = sorted(analytics.calls_by_day.items())
        for i, (day, count) in enumerate(sorted_days, 1):
            ws.cell(row=day_start_row + i, column=1, value=day).border = border
            ws.cell(row=day_start_row + i, column=2, value=count).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _export_summary_report(ws, analytics: Analytics):
        """Export summary report to worksheet."""
        report_lines = [
            "=== TELECOMMUNICATION DATA ANALYSIS ===",
            "",
            f"Total Calls: {analytics.total_calls}",
            f"Incoming Calls: {analytics.incoming_calls}",
            f"Outgoing Calls: {analytics.outgoing_calls}",
            f"Unique Phone Numbers: {analytics.unique_numbers}",
            f"Total Duration: {analytics.total_duration_minutes:.2f} minutes",
            f"Average Call Duration: {analytics.average_call_duration:.2f} minutes",
            "",
            "=== MOST FREQUENT NUMBERS ===",
        ]
        
        for i, (number, count) in enumerate(analytics.most_frequent_numbers, 1):
            report_lines.append(f"{i}. {number} ({count} calls)")
        
        report_lines.extend([
            "",
            "=== CALLS BY DAY ===",
        ])
        
        sorted_days = sorted(analytics.calls_by_day.items())
        for day, count in sorted_days:
            report_lines.append(f"{day}: {count} calls")
        
        # Write report
        for row, line in enumerate(report_lines, 1):
            cell = ws.cell(row=row, column=1, value=line)
            if line.startswith("==="):
                cell.font = Font(bold=True)
        
        # Set column width
        ws.column_dimensions['A'].width = 80
    
    @staticmethod
    def _export_common_contacts(ws, analytics: Analytics):
        """Export common contacts to worksheet."""
        if not analytics.common_contacts:
            ws.cell(row=1, column=1, value="No common contacts found across target numbers")
            return
        
        # Headers
        headers = ["Phone Number", "Target Numbers", "Count"]
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, contact in enumerate(analytics.common_contacts, 2):
            target_nums = ', '.join(contact['target_numbers'])
            ws.cell(row=row, column=1, value=contact['number']).border = border
            ws.cell(row=row, column=2, value=target_nums).border = border
            ws.cell(row=row, column=3, value=contact['count']).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


class EsubpoenaTollsTool:
    """Main application window using tkinter."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("📞 eSubpoena Tolls Tool")
        self.root.geometry("1400x900")
        
        # Data
        self.all_records = []  # All records from all files
        self.analytics = None
        self.current_target_number = None  # Currently selected target number
        self.filtered_records = []  # Records filtered by current target number
        
        # Setup UI
        self._setup_ui()
    
    def _setup_ui(self):
        """Setup the user interface."""
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        title_label = ttk.Label(header_frame, text="📞 eSubpoena Tolls Tool", 
                               font=('Arial', 24, 'bold'))
        title_label.pack(side=tk.LEFT)
        
        self.export_btn = ttk.Button(header_frame, text="Export to Excel", 
                                   command=self._export_to_excel, state='disabled')
        self.export_btn.pack(side=tk.RIGHT)
        
        # Notebook (tabs)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create tabs
        self._create_overview_tab()
        self._create_records_tab()
        self._create_analytics_tab()
        self._create_summary_tab()
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready to process files")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def _create_overview_tab(self):
        """Create the overview tab."""
        overview_frame = ttk.Frame(self.notebook)
        self.notebook.add(overview_frame, text="Overview")
        
        # Welcome message
        welcome_label = ttk.Label(overview_frame, text="Welcome to eSubpoena Tolls Tool",
                                 font=('Arial', 20, 'bold'))
        welcome_label.pack(pady=20)
        
        desc_label = ttk.Label(overview_frame, 
                              text="This tool processes telecommunication XML data and provides comprehensive analytics.",
                              font=('Arial', 12))
        desc_label.pack(pady=10)
        
        # File selection frame
        file_frame = ttk.LabelFrame(overview_frame, text="File Selection", padding="20")
        file_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # File selection buttons
        button_frame = ttk.Frame(file_frame)
        button_frame.pack(pady=20)
        
        self.file_btn = ttk.Button(button_frame, text="📁 Select XML Files", 
                                 command=self._select_files, style='Large.TButton')
        self.file_btn.pack(side=tk.LEFT, padx=5)
        
        self.clear_btn = ttk.Button(button_frame, text="🗑️ Clear All", 
                                  command=self._clear_all, style='Large.TButton')
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        # File info
        self.file_info_label = ttk.Label(file_frame, text="No files selected")
        self.file_info_label.pack()
        
        # Target number selector (for multiple files)
        self.target_frame = ttk.LabelFrame(file_frame, text="Target Phone Number", padding="10")
        self.target_frame.pack(fill=tk.X, pady=10)
        
        self.target_var = tk.StringVar()
        self.target_combo = ttk.Combobox(self.target_frame, textvariable=self.target_var, 
                                       state="readonly", width=30)
        self.target_combo.pack(side=tk.LEFT, padx=5)
        self.target_combo.bind('<<ComboboxSelected>>', self._on_target_changed)
        
        self.target_label = ttk.Label(self.target_frame, text="Select a target number to view its records")
        self.target_label.pack(side=tk.LEFT, padx=10)
        
        # Statistics frame
        self.stats_frame = ttk.LabelFrame(overview_frame, text="Quick Statistics", padding="20")
        self.stats_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # Statistics grid
        self.stats_grid = ttk.Frame(self.stats_frame)
        self.stats_grid.pack(fill=tk.X)
        
        # Configure grid columns
        for i in range(4):
            self.stats_grid.columnconfigure(i, weight=1)
    
    def _create_records_tab(self):
        """Create the call records tab."""
        records_frame = ttk.Frame(self.notebook)
        self.notebook.add(records_frame, text="Call Records")
        
        # Header
        header_frame = ttk.Frame(records_frame)
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.records_label = ttk.Label(header_frame, text="No call records loaded")
        self.records_label.pack(side=tk.LEFT)
        
        self.export_records_btn = ttk.Button(header_frame, text="Export to Excel", 
                                           command=self._export_to_excel, state='disabled')
        self.export_records_btn.pack(side=tk.RIGHT)
        
        # Table
        table_frame = ttk.Frame(records_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Create treeview
        columns = ("Direction", "Target Number", "Remote Number", "Normalized", "Date & Time", 
                  "Duration (min)", "Day of Week", "Source File")
        self.records_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=20)
        
        # Configure columns
        for col in columns:
            self.records_tree.heading(col, text=col)
            self.records_tree.column(col, width=120)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.records_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.records_tree.xview)
        self.records_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout
        self.records_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
    
    def _create_analytics_tab(self):
        """Create the analytics tab."""
        analytics_frame = ttk.Frame(self.notebook)
        self.notebook.add(analytics_frame, text="Analytics")
        
        # Title
        title_label = ttk.Label(analytics_frame, text="Analytics Dashboard",
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=10)
        
        # Statistics cards
        self.stats_cards_frame = ttk.Frame(analytics_frame)
        self.stats_cards_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Analytics content
        self.analytics_content = ttk.Frame(analytics_frame)
        self.analytics_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Most frequent numbers
        frequent_frame = ttk.LabelFrame(self.analytics_content, text="Most Frequent Numbers", padding="10")
        frequent_frame.pack(fill=tk.X, pady=5)
        
        columns = ("Rank", "Phone Number", "Call Count")
        self.frequent_tree = ttk.Treeview(frequent_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.frequent_tree.heading(col, text=col)
            self.frequent_tree.column(col, width=150)
        
        self.frequent_tree.pack(fill=tk.X)
        
        # Calls by day
        day_frame = ttk.LabelFrame(self.analytics_content, text="Calls by Day", padding="10")
        day_frame.pack(fill=tk.X, pady=5)
        
        columns = ("Date", "Call Count")
        self.day_tree = ttk.Treeview(day_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.day_tree.heading(col, text=col)
            self.day_tree.column(col, width=200)
        
        self.day_tree.pack(fill=tk.X)
        
        # Common contacts across target numbers
        common_frame = ttk.LabelFrame(self.analytics_content, text="Common Contacts Across Target Numbers", padding="10")
        common_frame.pack(fill=tk.X, pady=5)
        
        columns = ("Phone Number", "Target Numbers", "Count")
        self.common_tree = ttk.Treeview(common_frame, columns=columns, show='headings', height=8)
        
        for col in columns:
            self.common_tree.heading(col, text=col)
            self.common_tree.column(col, width=150)
        
        self.common_tree.pack(fill=tk.X)
        
        # No data message
        self.no_data_label = ttk.Label(analytics_frame, text="No analytics available. Please process an XML file first.",
                                      font=('Arial', 12))
        self.no_data_label.pack(pady=50)
    
    def _create_summary_tab(self):
        """Create the summary tab."""
        summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(summary_frame, text="Summary")
        
        # Header
        header_frame = ttk.Frame(summary_frame)
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(header_frame, text="Summary Report").pack(side=tk.LEFT)
        
        self.copy_btn = ttk.Button(header_frame, text="Copy to Clipboard", 
                                 command=self._copy_to_clipboard, state='disabled')
        self.copy_btn.pack(side=tk.RIGHT)
        
        # Summary text
        self.summary_text = scrolledtext.ScrolledText(summary_frame, height=30, font=('Consolas', 10))
        self.summary_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    def _select_files(self):
        """Select multiple XML files for processing."""
        file_paths = filedialog.askopenfilenames(
            title="Select XML Files",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
        
        if file_paths:
            self._process_files(file_paths)
    
    def _clear_all(self):
        """Clear all loaded data."""
        self.all_records = []
        self.filtered_records = []
        self.analytics = None
        self.current_target_number = None
        
        # Update UI
        self._update_records_table()
        self._update_analytics()
        self._update_summary()
        self._update_statistics()
        self._update_target_selector()
        
        # Disable export buttons
        self.export_btn.config(state='disabled')
        self.export_records_btn.config(state='disabled')
        self.copy_btn.config(state='disabled')
        
        # Update labels
        self.file_info_label.config(text="No files selected")
        self.target_label.config(text="Select a target number to view its records")
        
        # Switch to overview tab
        self.notebook.select(0)
        
        self.status_var.set("Ready to process files")
    
    def _on_target_changed(self, event=None):
        """Handle target number selection change."""
        selected = self.target_var.get()
        if selected and selected != "All Target Numbers":
            self.current_target_number = selected
            self.filtered_records = [r for r in self.all_records if r.target_number == selected]
        else:
            self.current_target_number = None
            self.filtered_records = self.all_records.copy()
        
        # Update UI with filtered records
        self._update_records_table()
        self._update_analytics()
        self._update_summary()
        self._update_statistics()
        
        # Update label
        if self.current_target_number:
            self.target_label.config(text=f"Showing records for: {self.current_target_number}")
        else:
            self.target_label.config(text="Showing all records")
    
    def _update_target_selector(self):
        """Update the target number selector dropdown."""
        if self.analytics and self.analytics.target_numbers:
            target_list = ["All Target Numbers"] + sorted(list(self.analytics.target_numbers))
            self.target_combo['values'] = target_list
            if target_list:
                self.target_combo.set(target_list[0])
        else:
            self.target_combo['values'] = []
            self.target_var.set("")
    
    def _select_file(self):
        """Select XML file for processing."""
        file_path = filedialog.askopenfilename(
            title="Select XML File",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
        
        if file_path:
            self._process_files([file_path])
    
    def _process_files(self, file_paths: List[str]):
        """Process multiple XML files."""
        self.status_var.set("Processing files...")
        file_names = [os.path.basename(f) for f in file_paths]
        self.file_info_label.config(text=f"Processing: {', '.join(file_names)}")
        
        # Start processing in a separate thread
        thread = threading.Thread(target=self._process_files_thread, args=(file_paths,))
        thread.daemon = True
        thread.start()
    
    def _process_files_thread(self, file_paths: List[str]):
        """Process multiple files in background thread."""
        try:
            all_records = []
            processed_files = []
            
            for file_path in file_paths:
                try:
                    records = XMLParser.parse_file(file_path)
                    all_records.extend(records)
                    processed_files.append(file_path)
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
                    continue
            
            if all_records:
                analytics = Analytics(all_records)
                # Update UI in main thread
                self.root.after(0, self._on_processing_finished, all_records, analytics, processed_files)
            else:
                self.root.after(0, self._on_processing_error, "No valid records found in any files")
            
        except Exception as e:
            self.root.after(0, self._on_processing_error, str(e))
    
    def _process_file(self, file_path: str):
        """Process the selected XML file."""
        self._process_files([file_path])
    
    def _process_file_thread(self, file_path: str):
        """Process file in background thread."""
        self._process_files_thread([file_path])
    
    def _on_processing_finished(self, records: List[CallRecord], analytics: Analytics, file_paths: List[str]):
        """Handle processing completion."""
        self.all_records = records
        self.analytics = analytics
        self.filtered_records = records.copy()  # Show all records initially
        
        file_names = [os.path.basename(f) for f in file_paths]
        self.status_var.set(f"Successfully processed {len(records)} call records from {len(file_paths)} files")
        self.file_info_label.config(text=f"Loaded: {', '.join(file_names)} ({len(records)} records)")
        
        # Update UI
        self._update_records_table()
        self._update_analytics()
        self._update_summary()
        self._update_statistics()
        self._update_target_selector()
        
        # Enable export buttons
        self.export_btn.config(state='normal')
        self.export_records_btn.config(state='normal')
        self.copy_btn.config(state='normal')
        
        # Switch to records tab
        self.notebook.select(1)
        
        messagebox.showinfo("Success", f"Successfully processed {len(records)} call records from {len(file_paths)} files!")
    
    def _on_processing_error(self, error_message: str):
        """Handle processing error."""
        self.status_var.set(f"Error: {error_message}")
        messagebox.showerror("Error", f"Failed to process file:\n{error_message}")
    
    def _update_records_table(self):
        """Update the records table."""
        # Clear existing items
        for item in self.records_tree.get_children():
            self.records_tree.delete(item)
        
        # Add new records
        for record in self.filtered_records:
            self.records_tree.insert('', 'end', values=(
                record.direction,
                record.target_number,
                record.remote_number,
                record.normalized_number,
                record.date_time,
                f"{record.duration_minutes:.2f}",
                record.day_of_week,
                record.source_file
            ))
        
        self.records_label.config(text=f"Showing {len(self.filtered_records)} call records")
    
    def _update_analytics(self):
        """Update the analytics tab."""
        if not self.analytics:
            return
        
        # Update statistics cards
        self._update_statistics_cards()
        
        # Update frequent numbers table
        for item in self.frequent_tree.get_children():
            self.frequent_tree.delete(item)
        
        for i, (number, count) in enumerate(self.analytics.most_frequent_numbers, 1):
            self.frequent_tree.insert('', 'end', values=(i, number, count))
        
        # Update calls by day table
        for item in self.day_tree.get_children():
            self.day_tree.delete(item)
        
        sorted_days = sorted(self.analytics.calls_by_day.items())
        for day, count in sorted_days:
            self.day_tree.insert('', 'end', values=(day, count))
        
        # Update common contacts table
        for item in self.common_tree.get_children():
            self.common_tree.delete(item)
        
        for contact in self.analytics.common_contacts:
            target_nums = ', '.join(contact['target_numbers'])
            self.common_tree.insert('', 'end', values=(
                contact['number'],
                target_nums,
                contact['count']
            ))
        
        # Show analytics content
        self.analytics_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        self.no_data_label.pack_forget()
    
    def _update_statistics_cards(self):
        """Update the statistics cards."""
        # Clear existing cards
        for widget in self.stats_cards_frame.winfo_children():
            widget.destroy()
        
        if not self.analytics:
            return
        
        # Create new cards
        stats = [
            ("Total Calls", str(self.analytics.total_calls), "📞"),
            ("Incoming", str(self.analytics.incoming_calls), "📥"),
            ("Outgoing", str(self.analytics.outgoing_calls), "📤"),
            ("Unique Numbers", str(self.analytics.unique_numbers), "👥"),
        ]
        
        for i, (title, value, icon) in enumerate(stats):
            card_frame = ttk.Frame(self.stats_cards_frame, relief=tk.RAISED, borderwidth=2)
            card_frame.grid(row=0, column=i, padx=5, pady=5, sticky=(tk.W, tk.E))
            
            icon_label = ttk.Label(card_frame, text=icon, font=('Arial', 16))
            icon_label.pack(pady=5)
            
            title_label = ttk.Label(card_frame, text=title, font=('Arial', 10))
            title_label.pack()
            
            value_label = ttk.Label(card_frame, text=value, font=('Arial', 16, 'bold'))
            value_label.pack(pady=5)
            
            self.stats_cards_frame.columnconfigure(i, weight=1)
    
    def _update_statistics(self):
        """Update the overview statistics."""
        if not self.analytics:
            return
        
        # Clear existing stats
        for widget in self.stats_grid.winfo_children():
            widget.destroy()
        
        # Add new stats
        stats = [
            ("Total Calls", str(self.analytics.total_calls)),
            ("Incoming", str(self.analytics.incoming_calls)),
            ("Outgoing", str(self.analytics.outgoing_calls)),
            ("Unique Numbers", str(self.analytics.unique_numbers)),
            ("Target Numbers", str(len(self.analytics.target_numbers))),
            ("Files Processed", str(len(self.analytics.files_processed))),
            ("Total Duration", f"{self.analytics.total_duration_minutes:.1f} min"),
            ("Avg Duration", f"{self.analytics.average_call_duration:.1f} min"),
        ]
        
        for i, (label, value) in enumerate(stats):
            row = i // 2
            col = (i % 2) * 2
            
            label_widget = ttk.Label(self.stats_grid, text=label, font=('Arial', 10, 'bold'))
            label_widget.grid(row=row, column=col, sticky=tk.W, padx=5, pady=2)
            
            value_widget = ttk.Label(self.stats_grid, text=value)
            value_widget.grid(row=row, column=col + 1, sticky=tk.W, padx=5, pady=2)
    
    def _update_summary(self):
        """Update the summary report."""
        if not self.analytics:
            return
        
        report_lines = [
            "=== TELECOMMUNICATION DATA ANALYSIS ===\n",
            f"Total Calls: {self.analytics.total_calls}",
            f"Incoming Calls: {self.analytics.incoming_calls}",
            f"Outgoing Calls: {self.analytics.outgoing_calls}",
            f"Unique Phone Numbers: {self.analytics.unique_numbers}",
            f"Target Numbers: {len(self.analytics.target_numbers)}",
            f"Files Processed: {len(self.analytics.files_processed)}",
            f"Total Duration: {self.analytics.total_duration_minutes:.2f} minutes",
            f"Average Call Duration: {self.analytics.average_call_duration:.2f} minutes\n",
            "=== TARGET NUMBERS ===",
        ]
        
        for target_num in sorted(self.analytics.target_numbers):
            target_records = [r for r in self.all_records if r.target_number == target_num]
            report_lines.append(f"• {target_num}: {len(target_records)} calls")
        
        report_lines.extend([
            "\n=== MOST FREQUENT NUMBERS ===",
        ])
        
        for i, (number, count) in enumerate(self.analytics.most_frequent_numbers, 1):
            report_lines.append(f"{i}. {number} ({count} calls)")
        
        if self.analytics.common_contacts:
            report_lines.extend([
                "\n=== COMMON CONTACTS ACROSS TARGET NUMBERS ===",
            ])
            
            for contact in self.analytics.common_contacts:
                target_nums = ', '.join(contact['target_numbers'])
                report_lines.append(f"• {contact['number']}: appears in {contact['count']} target numbers ({target_nums})")
        
        report_lines.extend([
            "\n=== CALLS BY DAY ===",
        ])
        
        sorted_days = sorted(self.analytics.calls_by_day.items())
        for day, count in sorted_days:
            report_lines.append(f"{day}: {count} calls")
        
        self.summary_text.delete(1.0, tk.END)
        self.summary_text.insert(1.0, "\n".join(report_lines))
    
    def _export_to_excel(self):
        """Export data to Excel."""
        if not self.all_records or not self.analytics:
            messagebox.showwarning("Warning", "No data to export")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                ExcelExporter.export_data(self.all_records, self.analytics, file_path)
                messagebox.showinfo("Success", f"Successfully exported to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed:\n{str(e)}")
    
    def _copy_to_clipboard(self):
        """Copy summary report to clipboard."""
        if self.summary_text.get(1.0, tk.END).strip():
            self.root.clipboard_clear()
            self.root.clipboard_append(self.summary_text.get(1.0, tk.END))
            messagebox.showinfo("Success", "Report copied to clipboard")


def main():
    """Main application entry point."""
    root = tk.Tk()
    app = EsubpoenaTollsTool(root)
    root.mainloop()


if __name__ == "__main__":
    main() 